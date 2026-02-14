import os
import time
import glob
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ───── CONFIGURATION ─────
DRIVER_PATH      = "msedgedriver.exe"
DEBUGGER_ADDRESS = "127.0.0.1:9222"
DOWNLOADS_DIR    = r"C:\Users\PC\Downloads"
MONTHLY_FOLDER   = r"D:\Kaam\Muneef 0\Monthly"
JUNE_SHEET      = "June 2025"
CHECKER_SHEET    = "All pizza stores monthly (June)"
COL_HEADER_ROW   = 4
ITEM_START_ROW   = 5  # checker items begin here

# Map web dropdown label → Excel filename
dropdown_to_file = {
    "Karachi Kabab Wala - Queen Street":    "Sheet2",
    "Pizza Karachi- Eglinton":       "Eglinton.xlsx",
    "Pizza Karachi -Heartland":      "pizza heartland.xlsx",
    "Karachi Kabab Wala":            "Kabab wala.xlsx",
    "Karachi Food Court":            "Karachi food court.xlsx",
    "Pizza Karachi Downtown TO":     "Pizza Downtown.xlsx",
    "Pizza Karachi- Highway Karahi": "Pizza highway.xlsx",
    "Pizza Karachi - Wonderland":    "pizza wonderland1.xlsx",
    "Pizza Karachi - Lebovic":       "Lebovic",
    "Pizza Karachi - Ajax":          "Pizza Ajax.xlsx",
    "Pizza Karachi - Lebovic":       "Lebovic.xlsx",
    "Pizza Karachi - Markham Rd":    "pizza markham.xlsx",
}

# ───── HELPERS ─────
def latest_file(folder, ext="*.xlsx"):
    files = glob.glob(os.path.join(folder, ext))
    return max(files, key=os.path.getctime)

def download_xlsx(driver, wait):
    wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Get this report')]"))).click()
    time.sleep(0.5)
    wait.until(EC.element_to_be_clickable((By.XPATH, "//li[contains(., 'Download')]"))).click()
    time.sleep(0.5)
    wait.until(EC.element_to_be_clickable((By.XPATH, "//li[contains(., 'XLSX')]"))).click()
    # give browser time to save
    time.sleep(8)
    return latest_file(DOWNLOADS_DIR)

# ───── SELENIUM SETUP ─────
opts = Options()
opts.use_chromium = True
opts.add_experimental_option("debuggerAddress", DEBUGGER_ADDRESS)
driver = webdriver.Edge(service=Service(DRIVER_PATH), options=opts)
wait = WebDriverWait(driver, 20)

try:
    # open the restaurant list
    btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[data-testid='radio-dropdown-selector-input']")))
    btn.click()
    time.sleep(0.5)

    for label, fname in dropdown_to_file.items():
        path = os.path.join(MONTHLY_FOLDER, fname)
        if not os.path.exists(path):
            print(f"❌ Workbook not found: {path}")
            continue
        print(f"\n▶ Processing {label} → {fname}")

        # select restaurant
        xpath = f"//div[@data-pw='radio-dropdown-selector']//label[.//span[text()='{label}']]"
        wait.until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'table[data-pw="report-table-data"]')))
        time.sleep(0.5)

        # download report
        downloaded = download_xlsx(driver, wait)
        print("  ↓ Downloaded", os.path.basename(downloaded))

        # load workbooks
        mb = load_workbook(path)
        ab = load_workbook(downloaded)
        ws_apr = ab.active

        # create or overwrite April sheet
        if JUNE_SHEET in mb.sheetnames:
            del mb[JUNE_SHEET]
        ws_new = mb.create_sheet(JUNE_SHEET)
        for r in ws_apr.iter_rows():
            for c in r:
                ws_new.cell(row=c.row, column=c.column).value = c.value

        # update checker
        ws_chk = mb[CHECKER_SHEET]
        # insert new column B
        ws_chk.insert_cols(2)
        ws_chk.cell(row=COL_HEADER_ROW, column=2).value = "April"

        # gather existing items
        existing = {ws_chk.cell(row=r, column=1).value for r in range(ITEM_START_ROW, ws_chk.max_row+1)}
        # gather april items from ws_new column A, rows 3..end except 'REPORT SUMMARY'
        apr_items = []
        for row in range(3, ws_new.max_row+1):
            val = ws_new.cell(row=row, column=1).value
            if val and "REPORT SUMMARY" not in str(val).upper():
                apr_items.append(val)

        # append missing items
        for item in apr_items:
            if item not in existing:
                new_r = ws_chk.max_row + 1
                ws_chk.cell(row=new_r, column=1).value = item

        # now fill formulas
        last_apr = ws_new.max_row
        col_A = get_column_letter(1)
        col_F = get_column_letter(6)
        for r in range(ITEM_START_ROW, ws_chk.max_row+1):
            itm = ws_chk.cell(row=r, column=1).value
            if not itm: continue
            formula = (
                f"=XLOOKUP({col_A}{r},"
                f"'{JUNE_SHEET}'!${col_A}$3:${col_A}${last_apr},"
                f"'{JUNE_SHEET}'!${col_F}$3:${col_F}${last_apr},"
                "0,0,1)"
            )
            ws_chk.cell(row=r, column=2).value = formula

        # save monthly workbook
        mb.save(path)
        print("  ✅ Updated", fname)

        # cleanup and reopen dropdown
        ab.close()
        os.remove(downloaded)
        btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[data-testid='radio-dropdown-selector-input']")))
        btn.click()
        time.sleep(0.5)

finally:
    driver.quit()
