from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import time

# ---------------- CONFIG ---------------- #
START_DATE = "2026-02-02"
END_DATE   = "2026-02-08"

RESTAURANTS = {
    "Chickentarian (Ajax)":                 "6d016df5-a637-53c8-9d64-56aeff7a20b9",
    "Chickentarian (Britannia)":            "a4851bd5-9160-5208-b556-92ba3426c6de",
    "Chickentarian (Lebovic)":              "41c579a9-1b2a-5c7f-9856-0038cbd1bb5b",
    "Chickentarian (Platinium Dr)":         "8b9deca8-cea4-5c1c-9f9f-817f57acb068",
    "Chickentarian (Toronto)":              "0a3b48a7-6941-5cc7-8547-53684bf48856",
    "Karachi Food Court":                   "6ceab5f3-2f2b-57bc-80e3-1a26dcd8c47a",
    "Karachi Highway Karahi":               "e89dff71-3cda-5432-a1be-ad745f4171fe",
    "Karachi Kabab Wala":                   "005936f1-ebc6-598d-9c36-42894a11bcaf",
    "Karachi Kabab Wala (Toronto)":         "75fabf19-0ce6-558b-a1a4-89f701f42e34",
    "Pizza Karachi (9661 Jane Street)":     "5c4046b8-537b-5b9f-9b93-647caf96869e",
    "Pizza Karachi (Dixie Rd)":             "fbd1e576-c136-5388-88e7-80544f954918",
    "Pizza Karachi (Eglinton Ave W)":       "2f84ac6c-8e14-54c3-a6c6-facc6bad14fa",
    "Pizza Karachi (Markham)":              "6384206d-c56d-5b5c-968d-fcb5e9df3f28",
    "Pizza Karachi (Mississauga)":          "5b2087ed-e5cc-594f-a2aa-b8750ed997cb",
    "Pizza Karachi (Queen)":                "88eefc7a-e42e-55b3-b966-d7564e14b457",
    "Pizza Karachi Ajax":                   "b556d092-3901-5bbd-b9d2-4245fda2d8dd",
    "Pizza Karachi Lebovic":                "114c0481-2f4e-5ad4-8d13-b7bbf01680b0",
    "Smash Ox Burger (Britannia)":          "9e84e415-4558-5ab7-a892-048014013440",
    "Smash OX Burger (Mississauga)":        "51aa0ed8-7918-50d4-91cd-2c6820465262",
    "Smash OX Burger (platinium Dr)":       "88563c02-d16f-5b65-8778-b970a7ad8b90",
    "Smash OX Burger (Vaughan)":            "2656a84c-a1ab-5651-9d15-68a2a2894a4a",
    "Smash Ox Burgers (toronto)":           "b1be03a8-44a2-5ea6-9bd9-57f659d29965",
    "Toronto Hot Chicken (Mississauga)":    "ee44351b-e6da-5552-a150-2fddd28d2b83"
}

BASE_URL = "https://merchants.ubereats.com/manager/payments"
# --------------------------------------- #

# Attach to existing Edge with remote debugging
edge_options = Options()
edge_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
service = Service("./msedgedriver.exe")
driver = webdriver.Edge(service=service, options=edge_options)
wait = WebDriverWait(driver, 30)

# Excel workbook setup
wb = Workbook()
ws = wb.active
ws.title = "Earnings"
headers = ["Restaurant", "Earnings", "Marketing"]
ws.append(headers)

# Scrape loop
for restaurant, uuid in RESTAURANTS.items():
    print(f"📊 Fetching for: {restaurant}")
    url = (
        f"{BASE_URL}"
        f"?restaurantUUID={uuid}"
        f"&start={START_DATE}"
        f"&end={END_DATE}"
        f"&rangeType=1"
    )
    driver.get(url)
    # short pause to let dynamic content render
    time.sleep(5)

    # Earnings: wait until a non-empty monolabel under the Earnings label appears
    earnings = "NOT FOUND"

    try:
        earnings_xpath = (
            "//li[@role='treeitem']"
            "[.//div[normalize-space()='Earnings']]"
            "//div[@data-baseweb='typo-monolabelmedium']"
        )
        earnings = driver.find_element(By.XPATH, earnings_xpath).text.strip()
    except Exception:
        pass

    # Marketing: may be missing, so use find_elements (no long wait)
    marketing = "NOT FOUND"

    try:
        marketing_xpath = (
            "//li[@role='treeitem']"
            "[.//div[normalize-space()='Marketing']]"
            "//div[@data-baseweb='typo-monolabelmedium']"
        )
        marketing_elements = driver.find_elements(By.XPATH, marketing_xpath)
        if marketing_elements:
            marketing = marketing_elements[0].text.strip()
    except Exception:
        pass


    print(f"   ➜ Earnings : {earnings}    |    Marketing : {marketing}")
    ws.append([restaurant, earnings, marketing])
    time.sleep(5)

# ========== Beautify Excel ==========

# header styles
header_font = Font(bold=True)
header_fill = PatternFill("solid", fgColor="CCFFCC")  # light green
center_align = Alignment(horizontal="center", vertical="center")
thin_side = Side(border_style="thin", color="000000")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

for col_cell in ws[1]:
    col_cell.font = header_font
    col_cell.fill = header_fill
    col_cell.alignment = center_align
    col_cell.border = thin_border

# auto column widths (simple heuristic)
col_widths = {}
for row in ws.rows:
    for cell in row:
        if cell.value is None:
            length = 0
        else:
            length = len(str(cell.value))
        col = cell.column_letter
        current = col_widths.get(col, 0)
        if length > current:
            col_widths[col] = length

for col, width in col_widths.items():
    # add a little padding
    ws.column_dimensions[col].width = width + 4

# apply alignment and border to data rows
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
    for cell in row:
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border

# create an Excel table for nicer UI (auto-detect range)
last_row = ws.max_row
table_ref = f"A1:C{last_row}"
tbl = Table(displayName="EarningsTable", ref=table_ref)
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tbl.tableStyleInfo = style
ws.add_table(tbl)

# Save file
filename = f"ubereats_earnings_{START_DATE}_to_{END_DATE}.xlsx"
wb.save(filename)
print(f"\n✅ Done! Saved file: {filename}")

driver.quit()
